import streamlit as st
import pandas as pd
import io
import re
import json
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP
from difflib import SequenceMatcher
from streamlit.components.v1 import html

# --- CONFIGURAÇÃO E PERSISTÊNCIA ---
st.set_page_config(page_title="Automatizador de Preços PRO", layout="wide")

# --- INJEÇÃO DE JAVASCRIPT ---
js_shortcuts = """
<script>
    const doc = window.parent.document;
    doc.addEventListener('keydown', function(e) {
        if (e.key === 'F1') {
            e.preventDefault();
            const inputs = doc.querySelectorAll('input[type="text"]');
            const searchInput = Array.from(inputs).find(el => el.placeholder.includes("Digite e pressione"));
            if (searchInput) searchInput.focus();
        }
        if (e.key === 'F5') {
            e.preventDefault();
            const inputs = doc.querySelectorAll('input[type="text"]');
            const searchInput = Array.from(inputs).find(el => el.placeholder.includes("Digite e pressione"));
            if (searchInput) {
                searchInput.value = "";
                searchInput.dispatchEvent(new Event('input', { bubbles: true }));
                searchInput.dispatchEvent(new Event('change', { bubbles: true }));
                searchInput.focus();
            }
        }
    });
</script>
"""
html(js_shortcuts, height=0)

DB_STORAGE = "master_database.csv"
USERS_STORAGE = "users_db.json"

if 'carrinho' not in st.session_state:
    st.session_state.carrinho = []

if 'replicar_data' not in st.session_state:
    st.session_state.replicar_data = None

def load_users():
    if not os.path.exists(USERS_STORAGE):
        return {
            "admin": {
                "password": "admin123", 
                "expiry": "2099-12-31", 
                "role": "admin",
                "permissions": {
                    "ver_estoque": True,
                    "ver_preco": True,
                    "ver_caixa": True,
                    "ver_quant": True,
                    "funcao_cotacao": True,
                    "funcao_vendas": True,
                    "busca_hibrido": True,
                    "busca_barras": True,
                    "busca_similaridade": True
                }
            }
        }
    with open(USERS_STORAGE, "r") as f: 
        return json.load(f)

def save_users(users):
    with open(USERS_STORAGE, "w") as f: 
        json.dump(users, f, indent=4)

# --- LOGIN ---
if 'autenticado' not in st.session_state:
    st.session_state.autenticado = False

def login():
    st.sidebar.title("🔐 Acesso Restrito")
    u = st.sidebar.text_input("Usuário")
    p = st.sidebar.text_input("Senha", type="password")
    if st.sidebar.button("Entrar"):
        users = load_users()
        if u in users and users[u]['password'] == p:
            exp = datetime.strptime(users[u]['expiry'], "%Y-%m-%d")
            if datetime.now() <= exp:
                st.session_state.autenticado = True
                st.session_state.user_role = users[u]['role']
                st.session_state.username = u
                st.rerun()
        st.sidebar.error("Acesso negado")

if not st.session_state.autenticado:
    login()
    st.stop()

# --- FUNÇÕES DE APOIO ---
def extra_round(valor):
    if pd.isna(valor): return valor
    return float(Decimal(str(valor)).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP))

def extract_all_barcodes(val):
    if pd.isna(val) or val is None: return []
    text = str(val).split('.')[0]
    return re.findall(r'\d{8,14}', text)

def clean_barcode_prefix(barcode):
    if len(barcode) > 8 and (barcode.startswith('0') or barcode.startswith('1')):
        return barcode[1:]
    return barcode

def similarity(a, b):
    return SequenceMatcher(None, str(a).upper().strip(), str(b).upper().strip()).ratio()

def extrair_detalhes(texto):
    return set(re.findall(r'(\d+\s?(?:g|gr|kg|l|lt|ml)\b)', str(texto).lower()))

# --- CARREGAMENTO DO BANCO ---
@st.cache_data
def get_master_db():
    if os.path.exists(DB_STORAGE):
        df = pd.read_csv(DB_STORAGE)
        df['Barcode'] = df['Barcode'].astype(str).str.replace(r'\.0$', '', regex=True)
        for col in ['Stock', 'caixa', 'QUANT']:
            if col not in df.columns: df[col] = 0
        return df
    return pd.DataFrame(columns=['Description', 'Barcode', 'Price', 'Stock', 'caixa', 'QUANT'])

# --- INTERFACE ---
tabs = ["📊 Cotação", "💰 Vendas", "⚙️ Gerenciar Banco"]
if st.session_state.user_role == "admin": tabs.append("👤 Usuários")
aba = st.sidebar.radio("Navegação", tabs)

# --- ABA 1: COTAÇÃO ---
if aba == "📊 Cotação":
    st.title("📊 Automatizador de Cotações")
    master_db = get_master_db()
    if master_db.empty: st.warning("Banco vazio.")
    
    # Resgate das permissões do usuário logado
    users = load_users()
    current_user_data = users.get(st.session_state.username, {})
    current_perms = current_user_data.get("permissions", {})
    
    # Filtra as regras de busca permitidas para este usuário
    modos_permitidos = []
    if current_perms.get("busca_hibrido", True) or st.session_state.user_role == "admin":
        modos_permitidos.append("Híbrido (Barras + Similaridade)")
    if current_perms.get("busca_barras", True) or st.session_state.user_role == "admin":
        modos_permitidos.append("Apenas Barras")
    if current_perms.get("busca_similaridade", True) or st.session_state.user_role == "admin":
        modos_permitidos.append("Apenas Similaridade")
        
    st.sidebar.header("Configurações")
    
    if modos_permitidos:
        modo = st.sidebar.selectbox("Regra de Busca:", modos_permitidos)
        
        ignorar_01 = st.sidebar.checkbox("Ignorar 0 ou 1 à esquerda no EAN", value=False)
        estoque_minimo = st.sidebar.number_input("Estoque mínimo para Barras:", min_value=0, value=1)
        discount = st.sidebar.number_input("Desconto (%)", 0.0)
        aplicar_arredondamento = st.sidebar.checkbox("Arredondar preços", value=True)

        st.sidebar.subheader("Opções de Download")
        opcao_salvamento = st.sidebar.radio(
            "Como deseja baixar o resultado?",
            ["Novo Arquivo (cotacao_corrigida.xlsx)", "Mesmo nome do arquivo original"]
        )

        target_file = st.file_uploader("Planilha de Destino", type=["xlsx"])
        if target_file:
            c1, c2 = st.columns(2)
            header_pos = c1.number_input("Linha do Cabeçalho:", 1, 100, 10)
            start_row = c2.number_input("Linha de início dos produtos:", 1, 1000, 11)
            t_df_view = pd.read_excel(target_file, header=header_pos-1)
            col1, col2, col3 = st.columns(3)
            desc_col = col1.selectbox("Coluna Descrição", t_df_view.columns)
            bar_col = col2.selectbox("Coluna Barras", t_df_view.columns)
            price_col = col3.selectbox("Coluna Preço", t_df_view.columns)

            if st.button("🚀 Processar e Preservar Formatação"):
                master_db['Stock'] = pd.to_numeric(master_db['Stock'], errors='coerce').fillna(0)
                product_map = dict(zip(master_db['Barcode'].astype(str), zip(master_db['Price'], master_db['Stock'])))
                
                if opcao_salvamento == "Mesmo nome do arquivo original":
                    output_name = target_file.name
                else:
                    output_name = "cotacao_corrigida.xlsx"

                target_file.seek(0)
                wb = openpyxl.load_workbook(target_file)
                ws = wb.active
                col_indices = {str(ws.cell(row=header_pos, column=i).value).strip(): i for i in range(1, ws.max_column + 1)}
                
                try:
                    d_idx = col_indices[desc_col.strip()]
                    b_idx = col_indices[bar_col.strip()]
                    p_idx = col_indices[price_col.strip()]
                    
                    for r in range(int(start_row), ws.max_row + 1):
                        d_val = ws.cell(row=r, column=d_idx).value
                        if not d_val or str(d_val).strip() == "": continue
                        
                        found_p = None
                        b_val = ws.cell(row=r, column=b_idx).value
                        
                        if "Barras" in modo or "Híbrido" in modo:
                            barcodes = extract_all_barcodes(b_val)
                            for b in barcodes:
                                if b in product_map:
                                    p_val, s_val = product_map[b]
                                    if s_val >= estoque_minimo:
                                        found_p = p_val
                                        break
                                elif ignorar_01:
                                    b_limpo = clean_barcode_prefix(b)
                                    if b_limpo in product_map:
                                        p_val, s_val = product_map[b_limpo]
                                        if s_val >= estoque_minimo:
                                            found_p = p_val
                                            break
                        
                        if found_p is None and ("Similaridade" in modo or "Híbrido" in modo):
                            best_sim = 0
                            d_det = extrair_detalhes(d_val)
                            for _, row_db in master_db.iterrows():
                                sim = similarity(d_val, row_db['Description'])
                                if sim >= 0.75 and d_det == extrair_detalhes(row_db['Description']):
                                    if sim > best_sim:
                                        best_sim = sim
                                        found_p = row_db['Price']
                        
                        if found_p is not None:
                            f_p = float(found_p) * (1 - (discount/100))
                            ws.cell(row=r, column=p_idx).value = extra_round(f_p) if aplicar_arredondamento else f_p

                    contador_final = 0
                    for r in range(int(start_row), ws.max_row + 1):
                        celula_preco = ws.cell(row=r, column=p_idx).value
                        try:
                            if celula_preco is not None and float(celula_preco) > 0:
                                contador_final += 1
                        except: continue

                    out = io.BytesIO()
                    wb.save(out)
                    st.success(f"Sucesso! Foram preenchidos **{contador_final}** itens com preços.")
                    st.download_button(f"📥 Baixar Planilha ({output_name})", out.getvalue(), output_name)
                    
                except Exception as e:
                    st.error(f"Erro ao processar: {e}")
    else:
        st.error("Seu usuário não possui permissão para utilizar nenhuma regra de busca. Contate o administrador.")

# --- ABA 2: VENDAS ---
elif aba == "💰 Vendas":
    st.title("💰 Consulta e Pré-Pedido")
    master_db = get_master_db()
    col_vendas_1, col_vendas_2 = st.columns([3, 1])
    
    with col_vendas_1:
        st.subheader("🔍 Busca e Adição Rápida")
        query = st.text_input("Pesquisar produto:", placeholder="Digite e pressione Enter", key="search_query")
        desc_geral = st.number_input("Desconto Padrão na Tabela (%)", 0.0, 100.0, 0.0)
        
        if query:
            search_terms = query.replace('%', ' ').split()
            mask = master_db['Description'].apply(lambda x: all(term.upper() in str(x).upper() for term in search_terms))
            results = master_db[mask].head(20).copy() 
            
            if not results.empty:
                if st.session_state.replicar_data:
                    with st.container():
                        st.warning("🔄 **Replicação de Família Detectada**")
                        rep = st.session_state.replicar_data
                        st.write(f"Deseja replicar Qtd: {rep['qtd']} e Preço: R$ {rep['preco']} para família '{rep['familia']}'?")
                        familia_results = results[results['Description'].str.contains(rep['familia'], case=False) & (results['Barcode'] != rep['ean'])]
                        escolha = st.radio("Como deseja replicar?", ["Replicar em todos", "Escolher específicos"], horizontal=True)
                        
                        if escolha == "Escolher específicos":
                            selecionados = st.multiselect("Itens:", familia_results['Description'].tolist(), default=familia_results['Description'].tolist())
                            if st.button("Confirmar Seleção"):
                                for _, rf in familia_results.iterrows():
                                    if rf['Description'] in selecionados:
                                        st.session_state.carrinho.append({"Descrição": rf['Description'], "EAN": rf['Barcode'], "Qtd": rep['qtd'], "Preço Unit": rep['preco'], "Total": extra_round(rep['preco'] * rep['qtd'])})
                                st.session_state.replicar_data = None
                                st.rerun()
                        else:
                            if st.button("Confirmar em Todos"):
                                for _, rf in familia_results.iterrows():
                                    st.session_state.carrinho.append({"Descrição": rf['Description'], "EAN": rf['Barcode'], "Qtd": rep['qtd'], "Preço Unit": rep['preco'], "Total": extra_round(rep['preco'] * rep['qtd'])})
                                st.session_state.replicar_data = None
                                st.rerun()
                
                for idx, row in results.iterrows():
                    p_sugestao = extra_round(float(row['Price']) * (1 - (desc_geral/100)))
                    with st.container():
                        c_desc, c_ean, c_preco, c_qtd, c_add = st.columns([3, 1.5, 1.2, 1, 0.5])
                        c_desc.write(f"**{row['Description']}**")
                        c_ean.write(f"`{row['Barcode']}`")
                        c_preco.write(f"R$ {p_sugestao}")
                        input_qtd = c_qtd.number_input("Qtd", 1, 1000, 1, key=f"qtd_{idx}")
                        if c_add.button("➕", key=f"btn_{idx}"):
                            st.session_state.carrinho.append({"Descrição": row['Description'], "EAN": row['Barcode'], "Qtd": input_qtd, "Preço Unit": p_sugestao, "Total": extra_round(p_sugestao * input_qtd)})
                            palavras = row['Description'].split()
                            if len(palavras) >= 2:
                                st.session_state.replicar_data = {"familia": f"{palavras[0]} {palavras[1]}", "qtd": input_qtd, "preco": p_sugestao, "ean": row['Barcode']}
                            st.rerun()

    with col_vendas_2:
        st.subheader("🛒 Seu Pedido")
        if st.session_state.carrinho:
            df_cart = pd.DataFrame(st.session_state.carrinho)
            for item in st.session_state.carrinho:
                st.write(f"{item['Qtd']}x {item['Descrição']} - **R$ {item['Total']}**")
            total_pedido = df_cart['Total'].sum()
            st.metric("Total Geral", f"R$ {extra_round(total_pedido)}")
            if st.button("🗑️ Limpar Tudo"):
                st.session_state.carrinho = []
                st.rerun()
            output_xlsx = io.BytesIO()
            wb_ped = openpyxl.Workbook()
            ws_ped = wb_ped.active
            ws_ped.append(["Descrição", "Código EAN", "Qtd", "Preço Unit.", "Total"])
            for item in st.session_state.carrinho:
                ws_ped.append([item['Descrição'], item['EAN'], item['Qtd'], item['Preço Unit'], item['Total']])
            wb_ped.save(output_xlsx)
            st.download_button("📥 Baixar Pedido", output_xlsx.getvalue(), "pedido.xlsx")
        else:
            st.info("Carrinho vazio.")

# --- ABA 3: GERENCIAR BANCO ---
elif aba == "⚙️ Gerenciar Banco":
    st.title("⚙️ Gerenciar Banco")
    f = st.file_uploader("Upload Banco (xlsx/csv)", type=["xlsx", "csv"])
    if f and st.button("💾 Salvar Banco"):
        df_file = pd.read_excel(f) if f.name.endswith('.xlsx') else pd.read_csv(f)
        
        pure_desc = [str(x).strip() if pd.notna(x) else "" for x in list(df_file.iloc[:, 0])] if len(df_file.columns) > 0 else [""] * len(df_file)
        pure_bar = [str(x).strip() if pd.notna(x) else "" for x in list(df_file.iloc[:, 1])] if len(df_file.columns) > 1 else [""] * len(df_file)
        pure_price = [float(x) if pd.notna(x) else 0.0 for x in list(df_file.iloc[:, 2])] if len(df_file.columns) > 2 else [0.0] * len(df_file)
        
        pure_stock = []
        for x in list(df_file.iloc[:, 3]):
            try: pure_stock.append(int(float(x)) if pd.notna(x) else 0)
            except: pure_stock.append(0)
            
        pure_caixa = []
        for x in list(df_file.iloc[:, 4]):
            try: pure_caixa.append(float(x) if pd.notna(x) else 0.0)
            except: pure_caixa.append(0.0)
            
        pure_quant = []
        for x in list(df_file.iloc[:, 5]):
            try: pure_quant.append(int(float(x)) if pd.notna(x) else 1)
            except: pure_quant.append(1)

        df_novo = pd.DataFrame({
            'Description': pure_desc,
            'Barcode': pure_bar,
            'Price': pure_price,
            'Stock': pure_stock,
            'caixa': pure_caixa,
            'QUANT': pure_quant
        })

        df_antigo = get_master_db()
        
        def get_num_only(v):
            m = re.search(r'\d+', str(v))
            return int(m.group()) if m else 1

        lista_final = []
        antigo_dict = {}
        if not df_antigo.empty:
            antigo_dict = dict(zip(df_antigo['Barcode'].astype(str), df_antigo['QUANT']))
            
        for idx, row in df_novo.iterrows():
            raw_desc = row['Description']
            raw_barcode = row['Barcode']
            raw_stock = row['Stock']
            raw_caixa = row['caixa']
            raw_quant = row['QUANT']
            
            barcode_str = re.sub(r'\D', '', str(raw_barcode).split('.')[0])
            if not barcode_str: continue
            
            if barcode_str in antigo_dict:
                quant_final = antigo_dict[barcode_str]
            else:
                quant_final = raw_quant
                
            divisor = get_num_only(quant_final)
            nova_caixa = pd.to_numeric(raw_caixa, errors='coerce')
            nova_caixa = float(nova_caixa) if not pd.isna(nova_caixa) else 0.0
            preco_calculado = nova_caixa / divisor if divisor > 0 else 0.0
            
            lista_final.append({
                'Description': raw_desc,
                'Barcode': barcode_str,
                'Price': preco_calculado,
                'Stock': int(raw_stock),
                'caixa': nova_caixa,
                'QUANT': quant_final
            })
            
        df_resultado = pd.DataFrame(lista_final)
        df_resultado.to_csv(DB_STORAGE, index=False)
        st.cache_data.clear()
        st.success("Banco Updated!")
        st.dataframe(df_resultado.head())

# --- ABA 4: USUÁRIOS (COM CONTROLE EXPANDIDO DE REGRAS DE BUSCA) ---
elif aba == "👤 Usuários":
    st.title("👤 Gestão de Usuários e Permissões")
    users = load_users()
    
    c_b1, c_b2 = st.columns(2)
    with c_b1:
        df_u = pd.DataFrame.from_dict(users, orient='index').reset_index()
        st.download_button("📥 Backup Usuários", df_u.to_csv(index=False).encode('utf-8'), "backup_users.csv")
    with c_b2:
        up_b = st.file_uploader("Restore Usuários", type=["csv"])
        if up_b and st.button("🔥 Restaurar"):
            df_r = pd.read_csv(up_b)
            new_u = {}
            for _, r in df_r.iterrows():
                try: perm = json.loads(r['permissions'].replace("'", '"'))
                except: perm = {"ver_estoque": True, "ver_preco": True, "ver_caixa": True, "ver_quant": True, "funcao_cotacao": True, "funcao_vendas": True, "busca_hibrido": True, "busca_barras": True, "busca_similaridade": True}
                new_u[str(r['index'])] = {"password": str(r['password']), "expiry": str(r['expiry']), "role": str(r['role']), "permissions": perm}
            save_users(new_u)
            st.rerun()
            
    st.divider()
    col_n, col_e = st.columns(2)
    
    with col_n:
        st.subheader("➕ Criar Novo Usuário")
        with st.form("Novo_Form"):
            nu = st.text_input("Nome do Usuário")
            np = st.text_input("Senha de Acesso")
            nd = st.number_input("Validade da Conta (dias)", 1, 365, 30)
            
            st.markdown("---")
            st.write("📋 **Permissões de Campos e Funções:**")
            p_estoque = st.checkbox("Acesso à Coluna: Estoque", value=True)
            p_preco = st.checkbox("Acesso à Coluna: Preço", value=True)
            p_caixa = st.checkbox("Acesso à Coluna: Caixa", value=True)
            p_quant = st.checkbox("Acesso à Coluna: Quantidade (Família)", value=True)
            p_cotacao = st.checkbox("Permitir Função: Processar Cotações", value=True)
            p_vendas = st.checkbox("Permitir Função: Consulta de Vendas / Pedido", value=True)
            
            st.write("🔍 **Permissões de Regras de Busca:**")
            p_hibrido = st.checkbox("Regra: Híbrido (Barras + Similaridade)", value=True)
            p_barras = st.checkbox("Regra: Apenas Barras", value=True)
            p_similaridade = st.checkbox("Regra: Apenas Similaridade", value=True)
            
            if st.form_submit_button("Criar Usuário com Permissões"):
                if nu and np:
                    users[nu] = {
                        "password": np,
                        "expiry": (datetime.now() + timedelta(days=nd)).strftime("%Y-%m-%d"),
                        "role": "user",
                        "permissions": {
                            "ver_estoque": p_estoque,
                            "ver_preco": p_preco,
                            "ver_caixa": p_caixa,
                            "ver_quant": p_quant,
                            "funcao_cotacao": p_cotacao,
                            "funcao_vendas": p_vendas,
                            "busca_hibrido": p_hibrido,
                            "busca_barras": p_barras,
                            "busca_similaridade": p_similaridade
                        }
                    }
                    save_users(users)
                    st.success(f"Usuário {nu} criado com sucesso!")
                    st.rerun()
                else:
                    st.error("Preencha o Usuário e Senha!")

    with col_e:
        st.subheader("📝 Editar Usuário Existente")
        u_sel = st.selectbox("Selecione o Usuário para Modificar", list(users.keys()))
        if u_sel:
            current_perms = users[u_sel].get("permissions", {
                "ver_estoque": True, "ver_preco": True, "ver_caixa": True, "ver_quant": True, "funcao_cotacao": True, "funcao_vendas": True,
                "busca_hibrido": True, "busca_barras": True, "busca_similaridade": True
            })
            
            ep = st.text_input("Senha Atual/Nova", value=users[u_sel]["password"])
            ex = st.text_input("Data de Expiração (AAAA-MM-DD)", value=users[u_sel]["expiry"])
            
            st.markdown("---")
            st.write(f"📋 **Modificar Permissões de '{u_sel}':**")
            
            up_estoque = st.checkbox("Acesso à Coluna: Estoque", value=current_perms.get("ver_estoque", True), key="e_est")
            up_preco = st.checkbox("Acesso à Coluna: Preço", value=current_perms.get("ver_preco", True), key="e_prc")
            up_caixa = st.checkbox("Acesso à Coluna: Caixa", value=current_perms.get("ver_caixa", True), key="e_cx")
            up_quant = st.checkbox("Acesso à Coluna: Quantidade (Família)", value=current_perms.get("ver_quant", True), key="e_qnt")
            up_cotacao = st.checkbox("Permitir Função: Processar Cotações", value=current_perms.get("funcao_cotacao", True), key="e_cot")
            up_vendas = st.checkbox("Permitir Função: Consulta de Vendas / Pedido", value=current_perms.get("funcao_vendas", True), key="e_vnd")
            
            st.write("🔍 **Modificar Regras de Busca:**")
            up_hibrido = st.checkbox("Regra: Híbrido (Barras + Similaridade)", value=current_perms.get("busca_hibrido", True), key="e_hib")
            up_barras = st.checkbox("Regra: Apenas Barras", value=current_perms.get("busca_barras", True), key="e_bar")
            up_similaridade = st.checkbox("Regra: Apenas Similaridade", value=current_perms.get("busca_similaridade", True), key="e_sim")
            
            if st.button("Salvar Alterações do Usuário"):
                users[u_sel].update({
                    "password": ep,
                    "expiry": ex,
                    "permissions": {
                        "ver_estoque": up_estoque,
                        "ver_preco": up_preco,
                        "ver_caixa": up_caixa,
                        "ver_quant": up_quant,
                        "funcao_cotacao": up_cotacao,
                        "funcao_vendas": up_vendas,
                        "busca_hibrido": up_hibrido,
                        "busca_barras": up_barras,
                        "busca_similaridade": up_similaridade
                    }
                })
                save_users(users)
                st.success(f"Permissões e dados de '{u_sel}' atualizados!")
                st.rerun()
                
            if u_sel != "admin" and st.button("❌ Excluir Usuário permanentemente"):
                del users[u_sel]
                save_users(users)
                st.rerun()
